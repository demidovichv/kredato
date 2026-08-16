#!/usr/bin/env python3
"""Обновляет открытые ставки в site/assets/data/rates.json
и актуализирует дату в site/index.html.

Источники:
- Основной: CBR xlsx-файлы по депозитам/кредитам
- Фолбэк: статические диапазоны, если источник недоступен
"""
import json
import re
import sys
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path

try:
    import requests
except ImportError as e:
    print(f"Missing dep: {e}. Run: pip install requests openpyxl")
    sys.exit(2)

REPO = Path(__file__).resolve().parents[1]
INDEX = REPO / "site" / "index.html"
RATES = REPO / "site" / "assets" / "data" / "rates.json"
STATUS = REPO / "site" / "assets" / "data" / "rates.status.json"

CBR_DEPOSITS = "https://www.cbr.ru/vfs/statistics/pdko/int_rat/deposits.xlsx"
CBR_LOANS = "https://www.cbr.ru/vfs/statistics/pdko/int_rat/loans_ind_new.xlsx"
CBR_MORTGAGE = "https://www.cbr.ru/vfs/statistics/BankSector/Mortgage/02_08_Rates_housing.xlsx"

HEADERS = {"User-Agent": "Mozilla/5.0"}


def fetch_xlsx(url: str, retries: int = 3, backoff: float = 1.5) -> bytes | None:
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
            return r.content
        except Exception as e:
            last_err = e
            print(f"  -> fetch failed {url} attempt={attempt}: {e}")
    print(f"  -> fetch exhausted retries for {url}: {last_err}")
    return None


def get_latest_row(ws, max_check: int = 200):
    """Return the last non-empty data row as (row_index, values)."""
    last = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_check, values_only=True)):
        if row[0] and any(v is not None for v in row[1:]):
            last = (i, row)
    return last


def extract_latest_deposits(data: bytes) -> tuple[float, float] | None:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"  -> openpyxl unavailable: {e}")
        return None

    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        print(f"  -> workbook load failed deposits: {e}")
        return None

    ws = wb.active or wb[wb.sheetnames[0]]
    idx, row = get_latest_row(ws, max_check=getattr(ws, "max_row", 200) or 200)
    if row is None:
        return None

    vals = []
    for cell in row[3:10]:  # columns D-J roughly
        if isinstance(cell, (int, float)):
            v = float(cell)
            if 3.0 <= v <= 20.0:
                vals.append(v)

    if len(vals) < 2:
        return None
    return round(min(vals), 1), round(max(vals), 1)


def extract_latest_loans(data: bytes) -> tuple[float, float] | None:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"  -> openpyxl unavailable: {e}")
        return None

    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        print(f"  -> workbook load failed loans: {e}")
        return None

    ws = wb.active or wb[wb.sheetnames[0]]
    idx, row = get_latest_row(ws, max_check=getattr(ws, "max_row", 200) or 200)
    if row is None:
        return None

    vals = []
    for cell in row[1:]:
        if isinstance(cell, (int, float)):
            v = float(cell)
            if 3.0 <= v <= 35.0:
                vals.append(v)

    if len(vals) < 2:
        return None
    return round(min(vals), 1), round(max(vals), 1)


def extract_latest_mortgage(data: bytes) -> tuple[float, float] | None:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        print(f"  -> openpyxl unavailable: {e}")
        return None

    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        print(f"  -> workbook load failed mortgage: {e}")
        return None

    ws = wb.active or wb[wb.sheetnames[0]]
    idx, row = get_latest_row(ws, max_check=getattr(ws, "max_row", 200) or 200)
    if row is None:
        return None

    vals = []
    for cell in row[1:]:
        if isinstance(cell, (int, float)):
            v = float(cell)
            if 1.0 <= v <= 25.0:
                vals.append(v)

    if len(vals) < 2:
        return None
    return round(min(vals), 1), round(max(vals), 1)


def fmt(n: float) -> str:
    return f"{n:.1f}".replace(".", ",")


def safe_range(rng: tuple[float, float] | None, fallback: tuple[float, float], label: str) -> tuple[float, float]:
    if rng:
        print(f"  -> {label}: {fmt(rng[0])}% – {fmt(rng[1])}%")
        return rng
    print(f"  -> {label}: fallback {fmt(fallback[0])}% – {fmt(fallback[1])}%")
    return fallback


def write_status(source: str, stale: bool = False, error: str | None = None):
    payload = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "source": source,
        "debug_source": source,
        "stale": bool(stale),
        "error": error,
    }
    STATUS.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Updated {STATUS}")


def write_rates_json(deposits: tuple[float, float], mortgage: tuple[float, float], loans: tuple[float, float]):
    today_str = date.today().strftime("%d.%m.%Y")
    data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "periods": {
            "6m": {
                "deposits": f"{fmt(deposits[0])}–{fmt(deposits[1])}%",
                "mortgage": f"{fmt(mortgage[0])}–{fmt(mortgage[1])}%",
                "loans": f"{fmt(loans[0])}–{fmt(loans[1])}%",
            },
            "1y": {
                "deposits": f"{fmt(deposits[0])}–{fmt(deposits[1])}%",
                "mortgage": f"{fmt(mortgage[0])}–{fmt(mortgage[1])}%",
                "loans": f"{fmt(loans[0])}–{fmt(loans[1])}%",
            },
            "3y": {
                "deposits": f"{fmt(deposits[0])}–{fmt(deposits[1])}%",
                "mortgage": f"{fmt(mortgage[0])}–{fmt(mortgage[1])}%",
                "loans": f"{fmt(loans[0])}–{fmt(loans[1])}%",
            },
        },
        "caption": f"Диапазоны — ориентир по открытым данным ЦБ и банков на {today_str}. Не является индивидуальной рекомендацией.",
        "debug_source": "cbr_xlsx",
    }
    RATES.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"Updated {RATES}")


def patch_index(on_date: str) -> bool:
    html = INDEX.read_text(encoding="utf-8")
    new_html, n = re.subn(
        r'(<h3>Ставки на )\d{2}\.\d{2}\.\d{4}( \(по открытым данным\)</h3>)',
        rf"\g<1>{on_date}\g<2>",
        html,
    )
    if n != 1:
        print(f"  -> patch_index failed: expected 1 substitution, got {n}")
        return False
    INDEX.write_text(new_html, encoding="utf-8")
    print(f"Patched {INDEX}")
    return True


def main() -> int:
    today_str = date.today().strftime("%d.%m.%Y")

    deposits_bytes = fetch_xlsx(CBR_DEPOSITS)
    loans_bytes = fetch_xlsx(CBR_LOANS)
    mortgage_bytes = fetch_xlsx(CBR_MORTGAGE)

    deposits = safe_range(
        extract_latest_deposits(deposits_bytes) if deposits_bytes else None,
        (12.8, 14.5),
        "deposits",
    )
    loans = safe_range(
        extract_latest_loans(loans_bytes) if loans_bytes else None,
        (15.0, 25.0),
        "loans",
    )
    mortgage = safe_range(
        extract_latest_mortgage(mortgage_bytes) if mortgage_bytes else None,
        deposits,
        "mortgage",
    )

    if deposits_bytes and loans_bytes and mortgage_bytes:
        write_status("cbr_xlsx")
    else:
        write_status("cbr_xlsx_fallback", error="one or more upstream fetches failed")

    write_rates_json(deposits, mortgage, loans)
    if not patch_index(today_str):
        return 1
    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
